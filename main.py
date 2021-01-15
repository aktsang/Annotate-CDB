    #!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue Jun 30 08:25:17 2020

@author: tsanga
"""


# PURPOSE:
#     Clone information was downloaded to a spreadsheet from the GENIE Construct Database.
#     Update the construct database spreadsheet in a new excel file with results from 
#     deep sequencing shown in the master summary file. 


import re
import os
import openpyxl

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles.colors import Color
from Bio import pairwise2

from cdb_globals import *


### file inputs

# # location of sequencing results file from Quantitative Genomics
# ms_file_loc = '/Users/tsanga/Documents/code/deepseq_update_cdb/testfiles/master_summary_QG_20190116_copy.xlsx'

# #location of CDB file
# cdb_database_loc = '/Users/tsanga/Documents/code/deepseq_update_cdb/testfiles/ConstructDB_combined_testing.xlsx'

# information on the master_summary from Quantitative Genomics
samplename_col = 'A'
plate_col = 'B'
well_col = 'C'
map2ref_var_col = 'V'
denovo_var_col = 'W'
common_var_col = 'X'
map2ref_seq_col = 'AC'
denovo_seq_col = 'AE'

# information on the CDB Lookup Table
constructname_col = 'A'
sequence_col = 'B'
construct_id_col = 'D'
categorynumber_col = 'E'
construct_loc_col = 'F'
comments_col = 'I'


# load lookup workbooks
master_summary_file = load_workbook(ms_file_loc)
ms= master_summary_file.active
msrows = ms.max_row + 1

cdb_file = load_workbook(cdb_database_loc)
cdb = cdb_file.active
cdbrows = cdb.max_row + 1


### Results workbook initialization
results_wb = Workbook()

# create worksheets for each type of result
results_ws0 = results_wb.active # the default active sheet
results_ws1 = results_wb.create_sheet("Combos Renamed")
results_ws2 = results_wb.create_sheet("Frameshifts")
results_ws3 = results_wb.create_sheet("Non-identical Sequences")
results_ws4 = results_wb.create_sheet("Mutation Count")
# results_ws5 = results_wb.create_sheet("Unidentified Entries")
results_ws6 = results_wb.create_sheet("Sanger-DeepSeq Comparison")
results_ws7 = results_wb.create_sheet("Partial Length Sequences")
results_ws8 = results_wb.create_sheet("No common_var")

# intial values of the excel results summary
results_ws1_row = 1
results_ws2_row = 1
results_ws3_row = 1
results_ws4_row = 4
# results_ws5_row = 1
results_ws6_row = 1
results_ws7_row = 1
results_ws8_row = 1

results_ws0.title = "Summary"
results_ws0['A1'].value = 'Worksheet name'
results_ws0['B1'].value = 'Description'
results_ws0['A2'].value = 'Combos Renamed'
results_ws0['B2'].value = 'Combo mutation names that contained underscores or dashes, which were corrected to spaces.'
results_ws0['A3'].value = 'Frameshifts'
results_ws0['B3'].value = 'Frameshifts found in the master summary data.'
results_ws0['A4'].value = 'Mapped and Denovo Mismatch'
results_ws0['B4'].value = 'Called mutations with mismatches between mapped and de novo assembly gene sequences.'
results_ws0['A5'].value = 'Mutation Count'
results_ws0['B5'].value = 'All non-SDM mutations found in the dataset.'
# results_ws0['A6'].value = 'Unidentified entries'
# results_ws0['B6'].value = 'Entries in the master summary not found in the CDB'
results_ws0['A7'].value = 'Sanger-Deepseq Comparison'
results_ws0['B7'].value = 'Clones with an existing Sanger Sequence compared to Deeq Seq'
results_ws0['A8'].value = 'Partial Length Sequences'
results_ws0['B8'].value = 'As title says'
results_ws0['A9'].value = 'No common_var'
results_ws0['B9'].value = 'If no mutation was call, the sequences were translated and searched for mutations, shown here. '

results_ws1['A1'].value = 'Old construct name'
results_ws1['B1'].value = 'New construct name'

results_ws2['A1'].value = 'Construct'
results_ws2['B1'].value = 'Frameshift identified'
results_ws2['C1'].value = 'Comment'

results_ws3['A1'].value = 'Construct'
results_ws3['B1'].value = 'common_var'
results_ws3['C1'].value = 'map2ref'
results_ws3['D1'].value = 'denovo'
results_ws3['E1'].value = 'map2ref sequence features vs denovo'

results_ws4['A1'].value = 'Total number of entries: '
results_ws4['A2'].value = 'Number of constructs updated: '
results_ws4['A4'].value = 'Mutation'
results_ws4['B4'].value = 'Number of constructs containing'

# results_ws5['A1'].value = 'Sample'
# results_ws5['B1'].value = 'Plate'
# results_ws5['C1'].value = 'Well'

results_ws6['A1'].value = 'Construct name'
results_ws6['B1'].value = 'Sanger Call'
results_ws6['C1'].value = 'map2ref call'
results_ws6['D1'].value = 'denovo call'
results_ws6['E1'].value = 'common call'
results_ws6['F1'].value = 'Mutations Identical?'
results_ws6['G1'].value = 'Clone Location'
results_ws6['H1'].value = 'Starting reads'

results_ws7['A1'].value = 'Construct name'
results_ws7['B1'].value = 'De Novo sequence'
results_ws7['C1'].value = 'Sequence Length'
results_ws7['D1'].value = 'Clone Location'
results_ws7['E1'].value = 'Comment'

results_ws8['A1'].value = 'constructname'
results_ws8['B1'].value = 'Plate Info'
results_ws8['C1'].value = 'map2ref analysis'
results_ws8['D1'].value = 'denovo analysis'

# the total number of constructs processed is added to the workbook at the end of the program. 

# newNames_wb = Workbook()
# newNames_ws = newNames_wb.active

### fix underscores and dashes in mutation list

from fix_combo_names import fixComboNames
for i in range(2, cdbrows):
    constructname = cdb[samplename_col + str(i)].value
    newconstructname = fixComboNames(constructname)
    
    if newconstructname is not None: # if a change was made by fixComboNames
        cdb[samplename_col + str(i)].value = newconstructname # updates the CDB name
        
        # The following code summarizes the changes in a spreadsheet. 
        results_ws1_row = results_ws1.max_row + 1
        results_ws1['A' + str(results_ws1_row)].value = constructname
        results_ws1['B' + str(results_ws1_row)].value = newconstructname

# results_ws1.save('/Users/tsanga/Documents/code/deepseq_update_cdb/testfiles/output/results_summary.xlsx')    
# results_ws1.close()
# newNames_wb.save('/Users/tsanga/Documents/code/deepseq_update_cdb/testfiles/output/Fixed_Combo_Names_summary.xlsx')    
# newNames_wb.close()

print('Fix combo names complete.')


### Update the ConstructDB list using the master summary sequencing data. 

platewellRegex = re.compile(r'A\d\d\d\d_[A-H]\d\d')
    
# loop through master summary 
for j in range(2, msrows):

    platename = ms[plate_col + str(j)].value # get plate name from row j
    wellname = ms[well_col + str(j)].value # get well name from row j
    
    searchtext = platename + '_' + wellname # the plate_well we're looking for, e.g. A0123_A01
    print('searchtext ' + searchtext)

    # Search the CDB list for the plate and well listing.
    for k in range(2, cdbrows):
        pwmatchfound = 0
        pwmatch = platewellRegex.search(str(cdb[construct_loc_col + str(k)].value)) 
        if pwmatch is not None:
            cdb_loc_compare = pwmatch.group() # extract the text of the matching pattern
            if cdb_loc_compare == searchtext: # compare the matched text to the desired plate and well
                pwmatchfound = 1
                cdbindex = k # save the row number of this entry
                break # exit cdb search loop when the entry is found. 
                
        # if k == cdbrows and pwmatchfound == 0: # plate and well info were not found in the construct database list.
        #     results_ws5_row += 1
        #     results_ws5['A' + str(results_ws5_row)].value = ms[samplename_col + str(j)].value
        #     results_ws5['B' + str(results_ws5_row)].value = ms[plate_col + str(j)].value
        #     results_ws5['C' + str(results_ws5_row)].value = ms[well_col + str(j)].value
            
    if pwmatchfound == 1:
        ### Name and Sequence update
        constructname = cdb[samplename_col + str(cdbindex)].value
        map2refvar = ms[map2ref_var_col + str(j)].value
        denovovar = ms[denovo_var_col + str(j)].value
        commonvar = ms[common_var_col + str(j)].value
        # read map2ref and denovo sequences
        map2refseq = ms[map2ref_seq_col + str(j)].value
        if map2refseq is not None:
            map2refseq = map2refseq.upper() # convert to all upper case
        denovoseq = ms[denovo_seq_col + str(j)].value
        if denovoseq is not None:
            denovoseq = denovoseq.upper() # convert to all upper case
        scaffold = str(cdb[categorynumber_col + str(cdbindex)].value)
    
        
        ### Check for existing Sanger sequence
        from sequence_funcs import checkSequence
        seqFound = checkSequence(cdb[sequence_col + str(cdbindex)].value)
        
        if seqFound == False: # no Sanger sequence found, so update the record
     
            
            
            if commonvar is not None: # For now, only work on records with a called mutation
            
                print(constructname + ', (m)' + map2refvar + ', (d)' + denovovar + ', (c)' + commonvar)
                
                ### update name
                from updateMutants import updateMutations
                newconstructname = updateMutations(constructname, map2refvar, map2refseq, denovovar, denovoseq, commonvar, scaffold)
                
                print(newconstructname)
                
                if newconstructname is not None:
                    cdb[samplename_col + str(cdbindex)].value = newconstructname[0]
                    
                    # TESTING ONLY - color the changed construct names yellow
                    cdb[samplename_col + str(cdbindex)].fill = PatternFill(fill_type="solid", fgColor="ffff00")
                    cdb[comments_col + str(cdbindex)].fill = PatternFill(fill_type="solid", fgColor="ffff00")
                    print('Name updated: ' + newconstructname[0])
                    constructs_updated += 1
                    
                ### udpate sequence
                    # pass to function
                    
                    from sequence_funcs import compareSeq
                    seqident = compareSeq(map2refseq, denovoseq)
                    if seqident == True:
                        cdb[sequence_col + str(cdbindex)].value = denovoseq # if assembled sequences are identical, plug in denovo sequence
                        # make the cell yellow
                        cdb[sequence_col + str(cdbindex)].fill = PatternFill(fill_type="solid", fgColor="ffff00")
                    else:
                        # common var called but sequences are different.
                        # make the cell blue
                        cdb[sequence_col + str(cdbindex)].fill = PatternFill(fill_type="solid", fgColor="00e5ff")
                        
                        # plug the denovoseq into the cdb record
                        cdb[sequence_col + str(cdbindex)].value = denovoseq
                        
                        # align the map2ref and denovo sequences
                        from sequence_funcs import simple_alignment
                        alignment_comment = simple_alignment(denovoseq, map2refseq)
                        
                        # add to results summary
                        results_ws3_row = results_ws3.max_row + 1
                        results_ws3['A' + str(results_ws3_row)].value = constructname
                        results_ws3['B' + str(results_ws3_row)].value = commonvar
                        results_ws3['C' + str(results_ws3_row)].value = map2refseq
                        results_ws3['D' + str(results_ws3_row)].value = denovoseq
                        results_ws3['E' + str(results_ws3_row)].value = alignment_comment
                        
                    ### Annotation functions
                    origComment = cdb[comments_col + str(cdbindex)].value
                    goodMutations = newconstructname[1]
                    extraMutations = newconstructname[2]
                    
                    from annotations import annotate
                    finalComment = annotate(origComment, goodMutations, extraMutations, map2refvar, denovovar, denovoseq, scaffold)
                    
                    from sequence_funcs import findFrameshift
                    frameshiftList = findFrameshift(denovovar)
                    if frameshiftList is not None:
                        if frameshiftList[0] == True:
                            finalComment += ' ' + 'frameshift-' + frameshiftList[1] + ', '
                            
                    cdb[comments_col + str(cdbindex)].value = finalComment
                    
                    if finalComment.find('partial-length') > -1:
                        results_ws7_row += 1
                        results_ws7['A' + str(results_ws7_row)].value = constructname
                        results_ws7['B' + str(results_ws7_row)].value = denovoseq
                        results_ws7['C' + str(results_ws7_row)].value = len(denovoseq)
                        results_ws7['D' + str(results_ws7_row)].value = searchtext
                        results_ws7['E' + str(results_ws7_row)].value = finalComment
                            
                    # ## other nucleotide polymorphisms comments
                    if denovoseq is not None:
                        from sequence_funcs import polymorphisms
                        other_polymorphisms = polymorphisms(scaffold, denovoseq, goodMutations)
                    # if other_polymorphisms is not None:
                    #     # finalComment += other_polymorphisms

            else: # commonvar is None
                if map2refseq != None and denovoseq != None:
                    from sequence_funcs import findMutations
                    mut_from_seq = findMutations(scaffold, map2refseq, denovoseq, searchtext)
                    if mut_from_seq is not None:
                        if len(mut_from_seq) > 0:
                            results_ws8_row += 1
                            results_ws8['A' + str(results_ws8_row)].value = constructname
                            results_ws8['B' + str(results_ws8_row)].value = searchtext
                            
                            map2refString = ''
                            denovoString = ''
                            for i in range(len(mut_from_seq[0])):
                                if i < len(mut_from_seq[0])-1:
                                    map2refString += mut_from_seq[0][i] + ' '
                                else: 
                                    map2refString += mut_from_seq[0][i]
                                    
                            for d in range(len(mut_from_seq[1])):
                                if d < len(mut_from_seq[1])-1:
                                    denovoString += mut_from_seq[1][d] + ' '
                                else:
                                    denovoString += mut_from_seq[1][d]
                            
                            results_ws8['C' + str(results_ws8_row)].value = map2refString
                            results_ws8['D' + str(results_ws8_row)].value = denovoString
            
                ### Frameshifts without a commonvar    
                origComment = cdb[comments_col + str(cdbindex)].value
            
                from sequence_funcs import findFrameshift
                frameshiftList = findFrameshift(denovovar)
                if frameshiftList is not None:
                    if frameshiftList[0] == True:
                        finalComment = origComment + '_DS: frameshift-' + frameshiftList[1] + ', ' # simply states the residue at which the frameshift starts.
                        cdb[comments_col + str(cdbindex)].fill = PatternFill(fill_type="solid", fgColor="ffff00")
                        constructs_updated += 1
                else:
                    finalComment = origComment
                        
                # If no commonvar exists, no need to look for problems besides frameshifts.
                
                # # from sequence_funcs import polymorphisms
                # if denovoseq is not None:
                #     other_polymorphisms = polymorphisms(scaffold, denovoseq, goodMutations)
                # if other_polymorphisms is not None:
                #     finalComment += other_polymorphisms
                
                cdb[comments_col + str(cdbindex)].value = finalComment
                        
            
            ### save frameshift results to results_ws2
            
            if frameshiftList is not None:
                results_ws2_row = results_ws2.max_row + 1
                results_ws2['A' + str(results_ws2_row)].value = constructname
                results_ws2['B' + str(results_ws2_row)].value = denovovar
                results_ws2['C' + str(results_ws2_row)].value = finalComment
                
            print('Comment: ' + finalComment)
            
            ## Other single nucleotide polymorphisms and insertion/deletions:
            from sequence_funcs import polymorphisms
            
            # if denovoseq is not None and goodMutations is not None:
            #     other_polymorphisms = polymorphisms(scaffold, denovoseq, goodMutations)
            # if other_polymorphisms is not None:
            #     finalComment += other_polymorphisms
        
        
        elif seqFound == True: # found a Sanger sequence, don't modify to CDB record
        
            # TESTING ONLY - color the unchanged construct records green
            cdb[samplename_col + str(cdbindex)].fill = PatternFill(fill_type="solid", fgColor="48ff00")
            cdb[sequence_col + str(cdbindex)].fill = PatternFill(fill_type="solid", fgColor="48ff00")
            
            constructname = cdb[samplename_col + str(cdbindex)].value
            # map2refvar = ms[map2ref_var_col + str(j)].value
            # denovovar = ms[denovo_var_col + str(j)].value
            commonvar = ms[common_var_col + str(j)].value
            
            results_ws6_row += 1
            
            
            results_ws6['A' + str(results_ws6_row)].value = constructname
            
            
            if map2refseq is not None and denovoseq is not None:
                from sequence_funcs import compareSanger
                sanger_muts = compareSanger(constructname, commonvar, map2refseq, denovoseq, scaffold, searchtext)
                if sanger_muts is not None:
                    results_ws6['B' + str(results_ws6_row)].value = sanger_muts[0]
                    if sanger_muts[1] == True:
                        results_ws6['F' + str(results_ws6_row)].value = 1
                        results_ws6['F' + str(results_ws6_row)].fill = PatternFill(fill_type = "solid", fgColor = "00ff00")
                        
                    else:
                        results_ws6['F' + str(results_ws6_row)].value = 0
                        results_ws6['F' + str(results_ws6_row)].fill = PatternFill(fill_type = "solid", fgColor = "ff9900")
        
                    if map2refvar is not None:
                        results_ws6['C' + str(results_ws6_row)].value = map2refvar
                    else:
                        results_ws6['C' + str(results_ws6_row)].value = sanger_muts[2]
                        #results_ws6['C' + str(results_ws6_row)].fill = PatternFill(fill_type = "solid", fgColor = "ffff00")
                        
                    if denovovar is not None:
                        results_ws6['D' + str(results_ws6_row)].value = denovovar
                    else:
                        results_ws6['D' + str(results_ws6_row)].value = sanger_muts[3]
                        #results_ws6['D' + str(results_ws6_row)].fill = PatternFill(fill_type = "solid", fgColor = "ffff00")
                        
                    if commonvar is not None:
                        results_ws6['E' + str(results_ws6_row)].value = commonvar
                    else:
                        results_ws6['E' + str(results_ws6_row)].value = ''
                        #results_ws6['E' + str(results_ws6_row)].fill = PatternFill(fill_type = "solid", fgColor = "ffff00")
                
                results_ws6['G' + str(results_ws6_row)].value = searchtext
                results_ws6['H' + str(results_ws6_row)].value = ms['D' + str(j)].value
                
            else:
                # no returns from deep sequencing.
                # just get the mutations from construct name to plug into results_summary
                from sequence_funcs import getSangerMutations
                sanger_muts = getSangerMutations(constructname, commonvar, map2refseq, denovoseq, scaffold, searchtext)
                if sanger_muts is not None:
                    results_ws6['B' + str(results_ws6_row)].value = sanger_muts
                results_ws6['G' + str(results_ws6_row)].value = searchtext
            
        
        
# tally dna mutations in dna_mutlist

# first traverse the dna_mutlist and add unique values to a new list
unique_muts = []
for l in range(0, len(dna_mutlist)):
    if dna_mutlist[l] not in unique_muts:
        unique_muts.append(dna_mutlist[l])

# add up the mutations
for m in range(0, len(unique_muts)):
    occurrences = dna_mutlist.count(unique_muts[m])
    results_ws4_row += 1
    results_ws4['A' + str(results_ws4_row)].value = unique_muts[m]
    results_ws4['B' + str(results_ws4_row)].value = occurrences

results_ws4['B1'].value = j-1 # fill in the number of constructs processed. 
results_ws4['B2'].value = constructs_updated # fill in the number of construct names changed

cdb_file.save(cdb_output_file)
cdb_file.close()
results_wb.save(results_summary_file)
results_wb.close()
master_summary_file.close()

print('The operation has completed.')
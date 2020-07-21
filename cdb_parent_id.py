#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Fri Jul 10 21:00:43 2020

@author: tsanga


"""

# This is a standalone script that can be run on the construct db spreadsheet. 
# Extracts the construct ID from the construct name and fills it into 
# column D (the Parent ID column). 

import re
import os
import openpyxl

cdbfileloc = '/Users/tsanga/Documents/Arthur_GENIE_Work/Tn5/DeepSeqResults/Analysis_scripting_and_testing/testfiles/ConstructDB_combined_testing.xlsx'

from openpyxl import Workbook
wb = Workbook() # create a new workbook
ws = wb.active # assign the active worksheet to ws object

from openpyxl import load_workbook
cdbfile = load_workbook(cdbfileloc) # open the workbook
cdbtable = cdbfile.active # assigning the active worksheet to cdbtable
rows = cdbtable.max_row + 1 # get the number of rows. +1 because counting starts at 0

cloneRegex = re.compile(r'\.\d+\.\d+') # i.e. looks for pattern like .421.5446 or .7.1

for i in range(2, rows):
    cellname= 'A' + str(i)
    constructname = cdbtable[cellname].value
    print(constructname)
    match = cloneRegex.search(constructname)
    if match is not None:
        constructid = match.group() # gives the match found, i.e. .421.5446
        constructid = constructid[1:] # remove the first dot to give id like 421.5446
        cdbtable['D' + str(i)].value = constructid
        print(constructid)
    

cdbfile.save('/Users/tsanga/Documents/Arthur_GENIE_Work/Tn5/DeepSeqResults/Analysis_scripting_and_testing/testfiles/ConstructDB_parentid_test.xlsx')
cdbfile.close